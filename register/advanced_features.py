from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Q
from datetime import datetime, timedelta
import csv
import io
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@login_required
def export_page(request):
    """Export data page view"""
    from register.models import File, FileRequest
    
    files = File.objects.all()
    requests = FileRequest.objects.all()
    
    context = {
        'departments': Department.objects.all(),
        'total_files': files.count(),
        'total_requests': requests.count(),
        'checked_out': files.filter(status='checked_out').count(),
        'overdue': files.filter(status='overdue').count(),
    }
    return render(request, 'register/export_data.html', context)

from register.models import File, FileMovement, FileRequest, Department


@login_required
def bulk_file_operation(request):
    """Handle bulk file operations"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    action = request.POST.get('action')
    file_ids = request.POST.getlist('file_ids')
    
    if not file_ids:
        return JsonResponse({'error': 'No files selected'}, status=400)
    
    count = 0
    if action == 'archive':
        for file_id in file_ids:
            try:
                file = File.objects.get(id=file_id)
                file.is_archived = True
                file.save()
                count += 1
            except File.DoesNotExist:
                pass
        messages.success(request, f'{count} file(s) archived.')
    
    elif action == 'delete':
        for file_id in file_ids:
            try:
                file = File.objects.get(id=file_id)
                file.delete()
                count += 1
            except File.DoesNotExist:
                pass
        messages.success(request, f'{count} file(s) deleted.')
    
    elif action == 'export':
        return JsonResponse({'redirect': f'/register/files/export/?ids={",".join(file_ids)}'})
    
    return JsonResponse({'success': True, 'count': count})


@login_required
def export_files(request):
    """Export files to CSV, PDF, or Excel"""
    file_ids = request.GET.get('ids', '').split(',')
    export_format = request.GET.get('format', 'csv')
    status_filter = request.GET.get('status', '')
    dept_filter = request.GET.get('department', '')
    
    files = File.objects.all()
    
    # Apply filters
    if file_ids and file_ids[0]:
        files = files.filter(id__in=file_ids)
    
    if status_filter:
        files = files.filter(status=status_filter)
    
    if dept_filter:
        files = files.filter(department_id=dept_filter)
    
    if export_format == 'csv':
        return export_files_csv(files)
    elif export_format == 'pdf':
        return export_files_pdf(files)
    elif export_format == 'xlsx':
        return export_files_excel(files)
    else:
        return JsonResponse({'error': 'Invalid format'}, status=400)


def export_files_csv(files):
    """Export files to CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(['Reference', 'Title', 'Status', 'Department', 'Current Holder', 'Created Date', 'Due Date'])
    
    # Data
    for file in files:
        writer.writerow([
            file.reference,
            file.title,
            file.get_status_display(),
            file.department.name if file.department else '',
            file.current_holder.username if file.current_holder else '',
            file.created_at.strftime('%Y-%m-%d'),
            file.due_date.strftime('%Y-%m-%d') if file.due_date else '',
        ])
    
    response = HttpResponse(output.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="files_{datetime.now().strftime("%Y%m%d")}.csv"'
    return response


def export_files_excel(files):
    """Export files to Excel with formatting"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse('Excel export not available. Install openpyxl.', status=500)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Files"
    
    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    headers = ['Reference', 'Title', 'Status', 'Department', 'Current Holder', 'Created Date', 'Due Date']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data
    for row, file in enumerate(files, 2):
        ws.cell(row=row, column=1, value=file.reference)
        ws.cell(row=row, column=2, value=file.title)
        ws.cell(row=row, column=3, value=file.get_status_display())
        ws.cell(row=row, column=4, value=file.department.name if file.department else '')
        ws.cell(row=row, column=5, value=file.current_holder.username if file.current_holder else '')
        ws.cell(row=row, column=6, value=file.created_at.strftime('%Y-%m-%d') if file.created_at else '')
        ws.cell(row=row, column=7, value=file.due_date.strftime('%Y-%m-%d') if file.due_date else '')
    
    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="files_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


def export_files_pdf(files):
    """Export files to PDF"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=0.5*inch, leftMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, spaceAfter=20)
    elements.append(Paragraph('File Report', title_style))
    elements.append(Spacer(1, 10))
    
    # Table data
    data = [['Reference', 'Title', 'Status', 'Department', 'Holder']]
    for file in files[:50]:  # Limit to 50 for PDF
        data.append([
            file.reference,
            file.title[:30] + '...' if len(file.title) > 30 else file.title,
            file.get_status_display(),
            file.department.name if file.department else '-',
            file.current_holder.username if file.current_holder else '-',
        ])
    
    # Create table
    table = Table(data, colWidths=[1*inch, 2.5*inch, 1*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="files_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response


@login_required
def export_requests(request):
    """Export file requests to CSV, PDF, or Excel"""
    export_format = request.GET.get('format', 'csv')
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    requests = FileRequest.objects.all().select_related('file', 'requesting_user')
    
    if status_filter:
        requests = requests.filter(status=status_filter)
    
    if date_from:
        requests = requests.filter(created_at__date__gte=date_from)
    
    if date_to:
        requests = requests.filter(created_at__date__lte=date_to)
    
    if export_format == 'csv':
        return export_requests_csv(requests)
    elif export_format == 'pdf':
        return export_requests_pdf(requests)
    elif export_format == 'xlsx':
        return export_requests_excel(requests)
    return JsonResponse({'error': 'Invalid format'}, status=400)


def export_requests_csv(requests):
    """Export requests to CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(['Reference', 'File', 'Requester', 'Status', 'Requested Date', 'Processed Date'])
    
    for req in requests:
        writer.writerow([
            f'REQ-{req.id:04d}',
            req.file.reference,
            req.requesting_user.username,
            req.get_status_display(),
            req.created_at.strftime('%Y-%m-%d'),
            req.processed_at.strftime('%Y-%m-%d') if req.processed_at else '',
        ])
    
    response = HttpResponse(output.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="requests_{datetime.now().strftime("%Y%m%d")}.csv"'
    return response


def export_requests_pdf(requests):
    """Export requests to PDF"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=0.5*inch, leftMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    elements.append(Paragraph('File Request Report', styles['Heading1']))
    elements.append(Spacer(1, 10))
    
    data = [['ID', 'File', 'Requester', 'Status', 'Date']]
    for req in requests[:50]:
        data.append([
            f'REQ-{req.id:04d}',
            req.file.reference[:20],
            req.requesting_user.username,
            req.get_status_display(),
            req.created_at.strftime('%Y-%m-%d'),
        ])
    
    table = Table(data, colWidths=[0.8*inch, 1.5*inch, 1.2*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="requests_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response


def export_requests_excel(requests):
    """Export requests to Excel"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse('Excel export not available', status=500)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Requests"
    
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    headers = ['ID', 'File Reference', 'Requester', 'Status', 'Requested Date', 'Processed Date']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row, req in enumerate(requests, 2):
        ws.cell(row=row, column=1, value=f'REQ-{req.id:04d}')
        ws.cell(row=row, column=2, value=req.file.reference)
        ws.cell(row=row, column=3, value=req.requesting_user.username)
        ws.cell(row=row, column=4, value=req.get_status_display())
        ws.cell(row=row, column=5, value=req.created_at.strftime('%Y-%m-%d') if req.created_at else '')
        ws.cell(row=row, column=6, value=req.processed_at.strftime('%Y-%m-%d') if req.processed_at else '')
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="requests_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
def advanced_filter(request):
    """Advanced filtering view for files"""
    query = request.GET.get('q', '')
    status = request.GET.get('status', '')
    department = request.GET.get('department', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    holder = request.GET.get('holder', '')
    
    files = File.objects.all()
    
    if query:
        files = files.filter(
            Q(reference__icontains=query) |
            Q(title__icontains=query) |
            Q(description__icontains=query)
        )
    
    if status:
        files = files.filter(status=status)
    
    if department:
        files = files.filter(department_id=department)
    
    if date_from:
        files = files.filter(created_at__date__gte=date_from)
    
    if date_to:
        files = files.filter(created_at__date__lte=date_to)
    
    if holder:
        files = files.filter(current_holder__username__icontains=holder)
    
    # Render results
    from django.template.loader import render_to_string
    html = render_to_string('register/includes/file_list_partial.html', {'files': files[:50]})
    
    return JsonResponse({'html': html, 'count': files.count()})


@login_required
def activity_timeline(request):
    """Activity timeline view with visual history"""
    days = int(request.GET.get('days', 7))
    start_date = timezone.now() - timedelta(days=days)
    
    # Get all activities
    movements = FileMovement.objects.filter(
        created_at__gte=start_date
    ).select_related('file', 'from_user', 'to_user').order_by('-created_at')[:100]
    
    # Group by date
    timeline = {}
    for m in movements:
        date_key = m.created_at.date().strftime('%Y-%m-%d')
        if date_key not in timeline:
            timeline[date_key] = []
        timeline[date_key].append({
            'time': m.created_at.strftime('%H:%M'),
            'file': m.file.reference,
            'file_uuid': m.file.uuid,
            'file_title': m.file.title,
            'action': m.get_action_display(),
            'from_user': m.from_user.username if m.from_user else '',
            'to_user': m.to_user.username if m.to_user else '',
            'action_type': m.action,
        })
    
    context = {
        'timeline': timeline,
        'days': days,
    }
    return render(request, 'register/activity_timeline.html', context)


@login_required
def search_files(request):
    """Global file search with autocomplete"""
    query = request.GET.get('q', '')
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    files = File.objects.filter(
        Q(reference__icontains=query) |
        Q(title__icontains=query)
    )[:10]
    
    results = []
    for file in files:
        results.append({
            'id': file.id,
            'uuid': str(file.uuid),
            'reference': file.reference,
            'title': file.title,
            'status': file.status,
            'url': f'/register/files/{file.uuid}/'
        })
    
    return JsonResponse({'results': results})


@login_required
def file_comparison(request, file_uuid, v1_id, v2_id):
    """Compare two file versions"""
    from register.models import FileVersion
    
    file = get_object_or_404(File, uuid=file_uuid)
    version1 = get_object_or_404(FileVersion, file=file, id=v1_id)
    version2 = get_object_or_404(FileVersion, file=file, id=v2_id)
    
    # Simple text comparison for notes
    diff = []
    if version1.notes and version2.notes:
        lines1 = version1.notes.split('\n')
        lines2 = version2.notes.split('\n')
        
        for i, line in enumerate(lines1):
            if i < len(lines2):
                if line != lines2[i]:
                    diff.append({
                        'line': i + 1,
                        'v1': line,
                        'v2': lines2[i],
                        'type': 'changed'
                    })
            else:
                diff.append({'line': i + 1, 'v1': line, 'v2': '', 'type': 'removed'})
        
        for i, line in enumerate(lines2[len(lines1):], start=len(lines1)):
            diff.append({'line': i + 1, 'v1': '', 'v2': line, 'type': 'added'})
    
    context = {
        'file': file,
        'version1': version1,
        'version2': version2,
        'diff': diff,
    }
    return render(request, 'register/file_comparison.html', context)


from django.shortcuts import get_object_or_404